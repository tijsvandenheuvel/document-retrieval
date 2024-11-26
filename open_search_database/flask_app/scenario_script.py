import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from db_opensearch import search_by_keyword, search_by_vector, generate_embeddings
import json

def load_queries():
    try:
        file_name = "../documents/20241029_Scenario Update.xlsx"
        query_column_name = "vraag"
        df = pd.read_excel(file_name)

        if df.empty or query_column_name not in df.columns:
            return []

        return df[query_column_name].dropna().tolist()
    
    except Exception as e:
        return {"error": str(e)}, 500

def get_queries():
    queries = load_queries()
    results = []
    for query in queries:
        results.append({
            "query": query,
            "keyword_search_results": []
        })

    return results

def load_results():
    try:
        file_name = "../documents/20241029_Scenario Update.xlsx"
        query_column_name = "vraag"
        answer_column_name = "mogelijk antwoord"
        df = pd.read_excel(file_name)

        if df.empty or query_column_name not in df.columns:
            return []

        queries = df[query_column_name].dropna().tolist()
        
        answers = df[answer_column_name].fillna("").tolist()
        # print(answers)
        
        results = []
        structured_results = []
        
        for query, answer in zip(queries, answers):
            k_search_results = search_by_keyword(query)
            k_titles = [hit["_source"]["file_path"].split('/')[-1] for hit in k_search_results["hits"]["hits"]]
            
            query_vector = generate_embeddings(query)
            v_search_results = search_by_vector(query_vector)
            v_titles = [hit["_source"]["file_path"].split('/')[-1] for hit in v_search_results["hits"]["hits"]]
            
            # k_titles = []
            # v_titles = []

            # Append the query and search results to the results list
            results.append({
                "query": query,
                "keyword_search_results": json.dumps(k_titles),
                "vector_search_results": json.dumps(v_titles)
            })
            
            # Ensure exactly 10 results for each (pad or truncate if necessary)
            k_titles = k_titles[:10] + [""] * (10 - len(k_titles))  # Truncate or pad to 10
            v_titles = v_titles[:10] + [""] * (10 - len(v_titles))  # Truncate or pad to 10

            # Add the query row
            structured_results.append({
                "Query": query,
                "search by keyword": k_titles[0] if k_titles else "",
                "search by vector": v_titles[0] if v_titles else "",
            })

            if answer:
                structured_results.append({
                    "Query": answer,
                    "search by keyword": k_titles[1] if 1 < len(k_titles) else "",
                    "search by vector": v_titles[1] if 1 < len(v_titles) else "",
                })
                for i in range(2, 10):
                    structured_results.append({
                        "Query": "",
                        "search by keyword": k_titles[i] if i < len(k_titles) else "",
                        "search by vector": v_titles[i] if i < len(v_titles) else "",
                    })
            else:
                for i in range(1, 10):
                    structured_results.append({
                        "Query": "",
                        "search by keyword": k_titles[i] if i < len(k_titles) else "",
                        "search by vector": v_titles[i] if i < len(v_titles) else "",
                    })

        # Convert transformed results to a pandas DataFrame
        df = pd.DataFrame(structured_results)

        write_to_file(df)

        return results

    except Exception as e:
        return {"error": str(e)}, 500
    
    
    
def write_to_file(df):
    # Write the DataFrame to an Excel file
    output_file = "../documents/20241029_Scenario_Results.xlsx"
    df.to_excel(output_file, index=False)
    
            # Apply styling using openpyxl
    workbook = load_workbook(output_file)
    sheet = workbook.active

    # Define zebra pattern fills
    fill1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    fill2 = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")  # Light gray

    current_fill = fill1
    previous_query = None

    # Apply the zebra pattern
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        query_cell = row[0]  # Query column          
        if query_cell.value :  # New query found
            if query_cell.value[0] != '(':
                current_fill = fill1 if current_fill == fill2 else fill2  # Alternate fills
                previous_query = query_cell.value

        # Apply the current fill to the entire row
        for cell in row:
            cell.fill = current_fill

    # Save the styled workbook
    workbook.save(output_file)

    print(f"Results successfully written to {output_file}")